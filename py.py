import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import warnings
import os

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl.styles.stylesheet')

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Формирование мультипаллетов Eco2b | Ozon")
        self.geometry("600x400")

        icon = tk.PhotoImage(file="items/icon.png")
        self.iconphoto(False, icon)

        bottom_frame = tk.Frame(self)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)

        self.progress_label = tk.Label(bottom_frame, text="", anchor="center")
        self.progress_label.pack()

        self.progress = ttk.Progressbar(bottom_frame, orient="horizontal", length=500, mode="determinate")
        self.progress.pack(pady=2)
        self.progress["maximum"] = 100

        self.ozon_data = {}        # Артикул -> кол-во
        self.dimensions_data = {}  # Артикул -> размеры
        self.all_base = {}         # Артикул -> всё вместе

        tk.Button(self, text="Загрузить накладную с OZON", command=self.load_ozon_file).pack(pady=5)

        frame = tk.Frame(self)
        frame.pack(pady=5)

        tk.Label(frame, text="Выберите машину:").pack(side=tk.LEFT, padx=(0,5))

        # Словарь с машинами и их максимальной высотой
        self.vehicles = {
            "671 Isuzu Elf (водитель Караулов)": 175,
            "696 Fusu (водитель Кузин)": 170,
        }

        # Создаем Combobox
        self.vehicle_combo = ttk.Combobox(frame, values=list(self.vehicles.keys()), state="readonly")
        self.vehicle_combo.current(0)  # Выбираем первую по умолчанию
        self.vehicle_combo.pack(side=tk.LEFT)

        # Поле для отображения высоты (только для показа, можно сделать disabled)
        tk.Label(frame, text="Макс. высота палеты (см):").pack(side=tk.LEFT, padx=(15,5))
        self.height_entry = tk.Entry(frame, width=5)
        self.height_entry.pack(side=tk.LEFT)
        self.height_entry.insert(0, str(self.vehicles[self.vehicle_combo.get()]))

        # При выборе машины обновляем высоту палеты
        self.vehicle_combo.bind("<<ComboboxSelected>>", self.on_vehicle_selected)

        tk.Button(self, text="Рассчитать палеты", command=self.calculate_pallets).pack(pady=10)

        # Автоматическая загрузка файла с размерами при запуске
        self.after(100, self.load_dimensions_file)

    def on_vehicle_selected(self, event=None):
        vehicle = self.vehicle_combo.get()
        height = self.vehicles.get(vehicle, 205)
        self.height_entry.delete(0, tk.END)
        self.height_entry.insert(0, str(height))

    def load_ozon_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not path:
            return
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            article = str(row[3].value).strip()
            quantity = row[5].value
            if article and quantity:
                self.ozon_data[article] = int(quantity)
        messagebox.showinfo("Успех", "Файл накладной загружен.")
        self.merge_data()

    def load_dimensions_file(self):
        path = os.path.join(os.path.dirname(__file__), "items/размеры.xlsx")
        if not os.path.exists(path):
            messagebox.showerror("Ошибка", f"Файл '{path}' не найден.\nПоложите его рядом с программой.")
            self.destroy()
            return
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл размеров:\n{e}")
            self.destroy()
            return

        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            if any(cell.value is None for cell in row[:6]):
                continue
            article = str(row[0].value).strip()
            try:
                width = float(row[1].value)
                length = float(row[2].value)
                height = float(row[3].value)
                weight = float(row[4].value)
                extra = str(row[5].value)  # значение из 6-го столбца, тип можно уточнить
            except ValueError:
                continue
            self.dimensions_data[article] = {
                'width': width,
                'length': length,
                'height': height,
                'weight': weight,
                'extra': extra
            }
        self.merge_data()  # Чтобы объединить с ozon_data, если оно уже загружено
        print(self.dimensions_data)

    def merge_data(self):
        self.all_base.clear()
        for article, qty in self.ozon_data.items():
            if article in self.dimensions_data:
                self.all_base[article] = {
                    **self.dimensions_data[article],
                    'quantity': qty
                }
    
    def calculate_pallets(self):
        try:
            max_height = float(self.height_entry.get() or 205)
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректную высоту палеты.")
            return

        base_pallet_width = 80
        base_pallet_length = 120
        tolerance = 15  # оверхэнг 5 см с каждой стороны

        pallet_width = base_pallet_width + tolerance
        pallet_length = base_pallet_length + tolerance

        from copy import deepcopy

        items = deepcopy(self.all_base)
        for item in items.values():
            item['remaining'] = item['quantity']

        packaging_priority = {'ж': 0, 'м': 1, 'ом': 2}

        def place_layer(layer_items):
            layer_plan = []
            grid = [[False] * int(pallet_length) for _ in range(int(pallet_width))]

            def fits(x, y, w, l):
                if x + w > pallet_width or y + l > pallet_length:
                    return False
                for i in range(int(x), int(x + w)):
                    for j in range(int(y), int(y + l)):
                        if i >= int(pallet_width) or j >= int(pallet_length) or grid[i][j]:
                            return False
                return True

            def occupy(x, y, w, l):
                for i in range(int(x), int(x + w)):
                    for j in range(int(y), int(y + l)):
                        if i < int(pallet_width) and j < int(pallet_length):
                            grid[i][j] = True

            sorted_articles = sorted(
                layer_items.items(),
                key=lambda x: (packaging_priority.get(x[1].get('extra', 'Ж'), 3), -x[1]['remaining'])
            )

            for article, data in sorted_articles:
                w1, l1 = data['width'], data['length']
                w2, l2 = l1, w1
                count = data['remaining']
                placed = 0

                for rotation in [(w1, l1), (w2, l2)]:
                    w, l = rotation
                    for x in range(int(pallet_width - w) + 1):
                        for y in range(int(pallet_length - l) + 1):
                            if fits(x, y, w, l) and count > 0:
                                occupy(x, y, w, l)
                                layer_plan.append({
                                    'article': article,
                                    'x': x,
                                    'y': y,
                                    'width': w,
                                    'length': l
                                })
                                placed += 1
                                count -= 1

                if placed > 0:
                    layer_items[article]['remaining'] -= placed
                    return layer_plan  # слой из одного артикула
            return []

        pallets = []
        pallet_number = 1

        while any(i['remaining'] > 0 for i in items.values()):
            current_pallet = []
            current_height = 0
            layer_num = 1
            total_weight = 0

            while True:
                layer_items = {
                    k: v for k, v in items.items() if v['remaining'] > 0
                }
                if not layer_items:
                    break

                layer_plan = place_layer(layer_items)
                if not layer_plan:
                    break

                max_layer_height = 0
                layer_summary = {}

                for box in layer_plan:
                    art = box['article']
                    h = items[art]['height']
                    w = items[art]['weight']
                    max_layer_height = max(max_layer_height, h)
                    if art not in layer_summary:
                        layer_summary[art] = {'count': 0, 'weight': 0}
                    layer_summary[art]['count'] += 1
                    layer_summary[art]['weight'] += w

                if current_height + max_layer_height > max_height:
                    for box in layer_plan:
                        items[box['article']]['remaining'] += 1
                    break

                for art, d in layer_summary.items():
                    current_pallet.append({
                        'article': art,
                        'count': d['count'],
                        'layer': layer_num
                    })
                    total_weight += d['weight']

                current_height += max_layer_height
                layer_num += 1

            pallets.append({
                'pallet_number': pallet_number,
                'layers': current_pallet,
                'weight': round(total_weight, 2),
                'height': round(current_height, 2)
            })
            pallet_number += 1
            
        output = []
        
        packaging_priority = {'ж': 0, 'м': 1, 'ом': 2}
        packaging_labels = {'ж': 'жёсткий слой', 'м': 'мягкий слой', 'ом': 'очень мягкий слой'}
        
        for pallet in pallets:
            output.append(f"Палета {pallet['pallet_number']}:")

            grouped_layers = {}
            for layer in pallet['layers']:
                grouped_layers.setdefault(layer['article'], []).append(layer)

            sorted_layers = []
            for article, layers in grouped_layers.items():
                extra = self.dimensions_data.get(article, {}).get('extra', 'ж').lower()
                priority = packaging_priority.get(extra, 3)
                sorted_layers.append((priority, article, extra, layers))

            sorted_layers.sort(key=lambda x: (x[0], x[1]))  # по жесткости, потом по артикулу

            new_layer_number = 1
            for _, article, extra, layers in sorted_layers:
                label = packaging_labels.get(extra, 'неизвестная жёсткость')
                for layer in layers:
                    line = f"  - Артикул: {layer['article']}, Кол-во: {layer['count']}, Слой: {new_layer_number}  #{label}"
                    output.append(line)
                    new_layer_number += 1

            info_line = f"  Общий вес: {pallet['weight']} кг | Занятая высота: {pallet['height']} см\n"
            output.append(info_line.strip())

        try:
            with open("результат_укладки.txt", "w", encoding="utf-8") as f:
                f.write("\n".join(output))
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{e}")
        else:
            messagebox.showinfo(tk.END, "Готово, результат сохранён в файл 'результат_укладки.txt'.\n")


if __name__ == "__main__":
    app = App()
    app.mainloop()
